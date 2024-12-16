from flask import Flask, render_template, request, redirect, flash, session, send_file, url_for # type: ignore
import xml.etree.ElementTree as ET 
import pandas as pd # type: ignore
import openpyxl # type: ignore
from fpdf import FPDF # type: ignore
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side # type: ignore
import sqlite3
from werkzeug.security import generate_password_hash
from werkzeug.security import check_password_hash
import json

app = Flask(__name__)
app.secret_key = 'clave_secreta_para_flask'

# Función para conectar y crear la base de datos
def init_db():
    conn = sqlite3.connect('app.db')  # Nombre del archivo de la base de datos
    cursor = conn.cursor()

    # Crear tabla de usuarios
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL
        )
    ''')

    # Crear tabla de facturas
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS facturas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            detalles TEXT NOT NULL,
            fecha TEXT NOT NULL,
            total REAL NOT NULL
        )
    ''')
    # Crear tabla de reportes
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS reportes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            nombre TEXT NOT NULL,
            fecha TEXT NOT NULL,
            total REAL NOT NULL,
            FOREIGN KEY (user_id) REFERENCES usuarios (id)
        )
    ''')
    conn.commit()
    conn.close()
    

# Llamar a la función para crear las tablas al inicio
init_db()

# Conectar a la base de datos SQLite
def get_db_connection():
    conn = sqlite3.connect('app.db')
    conn.row_factory = sqlite3.Row
    return conn


facturas_info = []

#conexion y uso de la base datos 
#login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        # Obtener el usuario de la base de datos
        conn = get_db_connection()
        user = conn.execute('SELECT * FROM usuarios WHERE email = ?', (email,)).fetchone()
        conn.close()
        
        # Verificar si el usuario existe y si la contraseña es correcta
        if user is None:
            flash("El correo electrónico no está registrado.")
            return redirect('/index')
        elif not check_password_hash(user['password'], password):
            flash("Contraseña incorrecta.")
            return redirect('/index')

        # Guardar el usuario en la sesión
        session['user_id'] = user['id']
        session['user_name'] = user['username']  # Guardar el nombre en la sesión
        session['user_email'] = user['email']    # Guardar el correo en la sesión
        
        flash("Has iniciado sesión correctamente.")
        return redirect('/upload')

    return render_template('index.html')


 # Cerrar sesión
@app.route('/logout')
def logout():
        # Eliminar el usuario de la sesión
        session.pop('user_id', None)
        flash("Has cerrado sesión correctamente.")
        return redirect('/index')

@app.route('/index')
def destino():

    return render_template('index.html')



#Registro
@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']

        # Encriptar la contraseña antes de guardarla
        hashed_password = generate_password_hash(password)

        # Guardar en la base de datos
        conn = get_db_connection()
        try:
            conn.execute('INSERT INTO usuarios (username, email, password)VALUES (?, ?, ?)'
                         , (username, email, hashed_password))
            conn.commit()
            flash("Registro exitoso. Ya puedes iniciar sesión.")
            return redirect('/registro')
        except sqlite3.IntegrityError:
            flash("El correo electrónico ya está registrado.")
            return redirect('/registro')
        finally:
            conn.close()
            #return redirect('/login')
    
    return render_template('registro.html')

#seccion factura
@app.route('/facturas')
def ver_facturas():
    
    conn = sqlite3.connect('app.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM facturas')
    facturas = cursor.fetchall()
    conn.close()

    return render_template('facturas.html', facturas=facturas)

@app.route('/crear_factura', methods=['POST'])
def crear_factura():
    if 'user_id' not in session:
        flash("Debes iniciar sesión para guardar reportes.")
        return redirect(url_for('login'))
    detalles = request.form['detalles']
    fecha = request.form['fecha']
    total = request.form['total']

    conn = sqlite3.connect('app.db')
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO facturas (detalles, fecha, total)
        VALUES (?, ?, ?)
    ''', (detalles, fecha, total))
    conn.commit()
    conn.close()

    return redirect('/facturas')

# Ruta para la página principal
@app.route('/')
def home():
    return redirect('/index')  # Redirige a la página de login

@app.route('/upload')
def upload():
    return render_template('upload.html')

#Ruta para página de inicio
#@app.route('/inicio')
#def inicio():
 #   return render_template('inicio.html')

#Ruta para sección en donde se almacenaran
@app.route('/guardar_reporte', methods=['POST'])
def guardar_reporte():
    if 'user_id' not in session:
        flash("Debes iniciar sesión para guardar reportes.")
        return redirect(url_for('login'))

    user_id = session['user_id']
    excel_filename = request.form.get('excel_filename')
    pdf_filename = request.form.get('pdf_filename')
    reporte_nombre = request.form.get('reporte_nombre')

    if not excel_filename or not pdf_filename or not reporte_nombre:
        flash("Información incompleta para guardar el reporte.")
        return redirect(url_for('upload'))

    # Calcular el total del reporte
    total = calcular_total_reporte(excel_filename)

    conn = get_db_connection()
    try:
        fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn.execute('''
            INSERT INTO reportes (user_id, nombre, fecha, total) 
            VALUES (?, ?, ?, ?)
        ''', (user_id, reporte_nombre, fecha, total))
        conn.commit()
        flash("Reporte guardado exitosamente.")
    except Exception as e:
        print(f"Error al guardar el reporte: {e}")
        flash("Hubo un error al guardar el reporte.")
    finally:
        conn.close()

    return redirect(url_for('reportes_anteriores'))

def calcular_total_reporte(excel_filename):
    try:
        df = pd.read_excel(excel_filename)
        total = df['Total con impuestos'].sum()  # Asegúrate de que esta columna exista en tu Excel
        return total
    except Exception as e:
        print(f"Error al calcular el total del reporte: {e}")
        return 0
#Ruta para sección en donde se almacenaran reportes anteriores
@app.route('/reportes.anteriores')
def reportes_anteriores():
    if 'user_id' not in session:
        flash("Debes iniciar sesión para guardar reportes.")
        return redirect(url_for('login'))

    user_id = session['user_id']
    nombre = request.args.get('nombre', '')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')

    conn = get_db_connection()
    query = '''SELECT * FROM reportes WHERE user_id = ?'''
    params = [user_id]

    if nombre:
        query += ' AND nombre LIKE ?'
        params.append(f"%{nombre}%")
    if fecha_inicio and fecha_fin:
        query += ' AND fecha BETWEEN ? AND ?'
        params.extend([fecha_inicio, fecha_fin])

    query += ' ORDER BY fecha DESC'
    reportes = conn.execute(query, params).fetchall()
    conn.close()

    return render_template('reportes.html', reportes=reportes, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, nombre=nombre)


# Ruta para subir los archivos XML y generar los reportes
@app.route('/upload', methods=['GET', 'POST'])
def upload_files():
    if 'user_id' not in session:
            print("Usuario no autenticado")
            flash("Debes iniciar sesión para acceder a esta página.")
            return redirect(url_for('login'))
     
    print("Iniciando proceso de carga de archivos")

    uploaded_files = request.files.getlist('xml_files')
    print(f"Número de archivos cargados: {len(uploaded_files)}")

    # Reiniciar facturas_info antes de procesar nuevos archivos
    global facturas_info
    facturas_info = []
    productos_info = []

    for file in uploaded_files:

        print(f"\nProcesando archivo: {file.filename}")
        try:
            # Leer el contenido del archivo para depuración
            content = file.read().decode('utf-8')
            print(f"Contenido del archivo {file.filename}:\n{content[:200]}...")  # Primeros 200 caracteres
            file.seek(0)  # Regresar al inicio del archivo

            tree = ET.parse(file)
            root = tree.getroot()
            print("XML parseado correctamente")

            # Extraer información básica de la autorización
            estado = root.find('estado').text if root.find('estado') is not None else None
            numero_autorizacion = root.find('numeroAutorizacion').text if root.find('numeroAutorizacion') is not None else None
            fecha_autorizacion = root.find('fechaAutorizacion').text if root.find('fechaAutorizacion') is not None else None
            ambiente = root.find('ambiente').text if root.find('ambiente') is not None else None

            print(f"Estado: {estado}")
            print(f"Número de autorización: {numero_autorizacion}")

            # Procesar el comprobante
            comprobante_element = root.find('comprobante')
            if comprobante_element is not None:
                comprobante_xml = comprobante_element.text.strip()
                if comprobante_xml.startswith('<?xml'):
                    comprobante_xml = comprobante_xml.split('?>', 1)[1].strip()
                print("Elemento comprobante encontrado y procesado")
            else:
                print(f"El archivo {file.filename} no contiene un elemento 'comprobante'")
                continue

            comprobante_root = ET.fromstring(comprobante_xml)
            print("Comprobante XML parseado correctamente")

            # Información tributaria
            info_tributaria = comprobante_root.find('infoTributaria')
            razon_social = info_tributaria.find('razonSocial').text if info_tributaria.find('razonSocial') is not None else "No especificado"
            nombre_comercial = info_tributaria.find('nombreComercial').text if info_tributaria.find('nombreComercial') is not None else "No especificado"
            ruc_vendedor = info_tributaria.find('ruc').text if info_tributaria is not None else None
            dir_Matriz = info_tributaria.find('dirMatriz').text if info_tributaria.find('dirMatriz') is not None else "No especificado"
            clave_acceso = info_tributaria.find('claveAcceso').text if info_tributaria is not None else None
            codigo_factura = info_tributaria.find('secuencial').text if info_tributaria is not None else None

            print(f"Información tributaria procesada - RUC: {ruc_vendedor}")

            # Información de la factura
            info_factura = comprobante_root.find('infoFactura')
            fecha_emision = info_factura.find('fechaEmision').text if info_factura is not None else None
            total_sin_impuestos = info_factura.find('totalSinImpuestos').text if info_factura is not None else None
            importe_total = info_factura.find('importeTotal').text if info_factura is not None else None
            forma_pago = info_factura.find('.//pagos/pago/formaPago').text if info_factura is not None else None  # Extraer forma de pago

            # Información del comprador
            razon_social_comprador = comprobante_root.find('.//razonSocialComprador').text if comprobante_root.find('.//razonSocialComprador') is not None else "No especificado"
            ruc_comprador = comprobante_root.find('.//identificacionComprador').text if comprobante_root.find('.//identificacionComprador') is not None else "No especificado"

            print(f"Información de factura procesada - Fecha: {fecha_emision}")

            # Procesar detalles y productos
            detalles = comprobante_root.find('detalles')
            if detalles is not None:
                print("Procesando detalles de productos")
                productos = detalles.findall('detalle')
                ivas = {"0%": 0, "5%": 0, "12%": 0, "15%": 0}
                total_factura = 0

                for producto in productos:
                    codigo = producto.find('codigoPrincipal').text if producto.find('codigoPrincipal') is not None else None
                    descripcion = producto.find('descripcion').text if producto.find('descripcion') is not None else None
                    cantidad = float(producto.find('cantidad').text) if producto.find('cantidad') is not None else 0.0
                    precio_unitario = float(producto.find('precioUnitario').text) if producto.find('precioUnitario') is not None else 0.0
                    precio_total_sin_impuesto = float(producto.find('precioTotalSinImpuesto').text) if producto.find('precioTotalSinImpuesto') is not None else 0.0

                    # Procesar impuestos del producto
                    impuesto_element = producto.find('impuestos/impuesto')
                    if impuesto_element is not None:
                        impuesto = float(impuesto_element.find('valor').text) if impuesto_element.find('valor') is not None else 0.0
                        tarifa = float(impuesto_element.find('tarifa').text) if impuesto_element.find('tarifa') is not None else 0.0
                        
                        # Sumar al IVA correspondiente
                        if tarifa == 0:
                            ivas["0%"] += impuesto
                        elif tarifa == 5:
                            ivas["5%"] += impuesto
                        elif tarifa == 12:
                            ivas["12%"] += impuesto
                        elif tarifa == 15:
                            ivas["15%"] += impuesto
                    else:
                        impuesto = 0.0
                        tarifa = 0.0

                    total_producto = precio_total_sin_impuesto + impuesto
                    total_factura += total_producto

                    productos_info.append({
                        'Codigo Factura': codigo_factura,
                        'Código': codigo,
                        'Descripción': descripcion,
                        'Cantidad': cantidad,
                        'Precio Unitario': precio_unitario,
                        'Precio Total Sin Impuesto': precio_total_sin_impuesto,
                        'IVA': impuesto,
                        'Total': total_producto
                    })
                    print(f"Producto procesado: {codigo} - {descripcion}")

            # Agregar información de la factura
            factura_info = {
                'Codigo Factura': codigo_factura,
                'Estado de la autorización': estado,
                'Fecha de autorización': fecha_autorizacion,
                'Ambiente': ambiente,
                'Razón Social comprador': razon_social_comprador,
                'RUC del Comprador': ruc_comprador,
                'Razón Social del Vendedor': razon_social,
                'Nombre Comercial': nombre_comercial,
                'dir Establecimiento': dir_Matriz,
                'RUC del Vendedor': ruc_vendedor,
                'Fecha de Emisión': fecha_emision,
                'IVA 0%': ivas["0%"],
                'IVA 5%': ivas["5%"],
                'IVA 12%': ivas["12%"],
                'IVA 15%': ivas["15%"],
                'Total sin impuestos': total_sin_impuestos,
                'Total con impuestos': importe_total,
                'Número de autorización': numero_autorizacion,
                'Clave de Acceso': clave_acceso,
                'Forma Pago': forma_pago
            }
            facturas_info.append(factura_info)
            print(f"Factura procesada: {codigo_factura} - {razon_social_comprador}")

        except ET.ParseError:
            print(f"Error al analizar el archivo {file.filename}: El archivo no es un XML válido.")
        except Exception as e:
            print(f"Error procesando el archivo {file.filename}: {str(e)}")

    print(f"Total de facturas procesadas: {len(facturas_info)}")
    print(f"Total de productos procesados: {len(productos_info)}")


    # Generar reporte en Excel
    df_facturas = pd.DataFrame(facturas_info)
    df_productos = pd.DataFrame(productos_info)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo_excel = f'reporte_facturas_{timestamp}.xlsx'
    

    writer = pd.ExcelWriter(archivo_excel, engine='openpyxl')

    # Guardar el DataFrame en Excel
    df_facturas.to_excel(writer, index=False, sheet_name='Reporte')
    df_productos.to_excel(writer, sheet_name='Productos', index=False)

    # Obtener la hoja de trabajo
    workbook = writer.book
    worksheet_facturas = workbook['Reporte']
    worksheet_productos = workbook['Productos']

    # Aplicar formato a la hoja
    for idx, col in enumerate(df_facturas.columns):
        # Establecer el ancho de la columna basado en la longitud máxima en la columna
        max_len = max(df_facturas[col].astype(str).map(len).max(), len(col))
        worksheet_facturas.column_dimensions[openpyxl.utils.get_column_letter(idx+1)].width = max_len + 2

    for idx, col in enumerate(df_productos.columns):
        # Establecer el ancho de la columna basado en la longitud máxima en la columna
        max_len = max(df_productos[col].astype(str).map(len).max(), len(col))
        worksheet_productos.column_dimensions[openpyxl.utils.get_column_letter(idx+1)].width = max_len + 2
    
    for worksheet in [worksheet_facturas, worksheet_productos]:
    # Aplicar estilo a la fila de encabezados
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')

    # Aplicar bordes a todas las celdas
        thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border

    # Aplicar filtro
    worksheet.auto_filter.ref = worksheet.dimensions

    # Congelar la primera fila
    worksheet.freeze_panes = 'A2'

    # Guardar el archivo
    writer.close()


   

    # Función para generar el PDF con diseño de factura

    def generar_pdf_facturas(facturas_info, productos_info=None):
        if productos_info is None:
            productos_info = []

        archivo_pdf = f'reporte_facturas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=20)

        for factura in facturas_info:
            pdf.add_page()

            # Añadir el cuadro principal de información
            pdf.set_xy(10, 10)
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(190, 10, "NO TIENE LOGO", border=1, ln=1, align='C')

            # Cuadro izquierdo: Información del vendedor
            pdf.set_xy(10, 20)
            pdf.set_font('Arial', '', 10)
            pdf.multi_cell(85, 8, f"""
            {factura['Nombre Comercial']}
            {factura['dir Establecimiento']}
            """, border=1)

            # Cuadro derecho: Información del emisor
            pdf.set_xy(100, 20)
            pdf.multi_cell(100, 8, f"""
            RUC: {factura['RUC del Vendedor']}
            FACTURA: {factura['Codigo Factura']}
            NÚMERO DE AUTORIZACIÓN: {factura['Número de autorización']}
            FECHA Y HORA DE AUTORIZACIÓN: {factura['Fecha de Emisión']}
            AMBIENTE: PRODUCCIÓN
            EMISIÓN: NORMAL
            CLAVE DE ACCESO: {factura['Clave de Acceso']}
            """, border=1, align='L')

            # Cuadro inferior: Información del comprador
            pdf.set_xy(10, 70)
            pdf.multi_cell(85, 8, f"""
            Razón Social/Nombres: 
            {factura['Razón Social comprador']}
            Identificación: {factura['RUC del Comprador']}
            Fecha: {factura['Fecha de Emisión']}
            """, border=1)

            # Tabla de productos
            pdf.set_xy(10, 130)
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(40, 8, "Código", 1)
            pdf.cell(70, 8, "Descripción", 1)
            pdf.cell(20, 8, "Cantidad", 1)
            pdf.cell(30, 8, "P. Unitario", 1)
            pdf.cell(30, 8, "Total", 1, 1)

            pdf.set_font('Arial', '', 10)
            productos_factura = [
                producto for producto in productos_info
                if producto['Codigo Factura'] == factura['Codigo Factura']
            ]
            for producto in productos_factura:
                # Posición inicial de la fila
                x = pdf.get_x()
                y = pdf.get_y()

                # Descripción (MultiCell)
                pdf.cell(40, 8, str(producto['Código']), 1, 0)  # Código
                pdf.multi_cell(70, 8, str(producto['Descripción']), 1)  # Descripción
                # Ajusta la posición para las demás columnas
                max_y = pdf.get_y()
                pdf.set_xy(x + 110, y)
                pdf.cell(20, max_y - y, str(producto['Cantidad']), 1, 0, 'C')  # Cantidad
                pdf.cell(30, max_y - y, f"${producto['Precio Unitario']:.2f}", 1, 0, 'C')  # Precio Unitario
                pdf.cell(30, max_y - y, f"${producto['Total']:.2f}", 1, 1, 'C')  # Total

            # Totales
            pdf.ln(5)
            pdf.cell(160, 8, "Subtotal:", 0, 0, 'R')
            pdf.cell(30, 8, f"${float(factura['Total sin impuestos']):.2f}", 1, 1, 'R')

            for iva_tipo in ['IVA 0%', 'IVA 5%', 'IVA 12%', 'IVA 15%']:
                if iva_tipo in factura and float(factura[iva_tipo]) > 0:
                    pdf.cell(160, 8, f"{iva_tipo}:", 0, 0, 'R')
                    pdf.cell(30, 8, f"${float(factura[iva_tipo]):.2f}", 1, 1, 'R')

            pdf.cell(160, 8, "Total con impuestos:", 0, 0, 'R')
            pdf.cell(30, 8, f"${float(factura['Total con impuestos']):.2f}", 1, 1, 'R')

            # Cuadro en la esquina inferior para forma de pago y total
            pdf.ln(4)
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(80, 10, "Detalles de Pago", border=1, ln=1, align='C')
            pdf.set_font('Arial', '', 10)
            # Diccionario de formas de pago
            formas_pago = {
                '1': 'Efectivo', '2': 'Cheque', '3': 'Tarjeta de crédito',
                '4': 'Tarjeta de débito', '5': 'Transferencia bancaria',
                '6': 'Dinero electrónico', '7': 'Crédito', '8': 'Pago anticipado',
                '9': 'Compensación',
                '01': 'Efectivo', '02': 'Cheque', '03': 'Tarjeta de crédito',
                '04': 'Tarjeta de débito', '05': 'Transferencia bancaria',
                '06': 'Dinero electrónico', '07': 'Crédito', '08': 'Pago anticipado',
                '09': 'Compensación', '10': 'Pago en especie', '11': 'Cesión de derechos',
                '12': 'Pago en especie o compensación', '13': 'Tarjeta prepago',
                '14': 'Pago con bonos', '15': 'Pago por servicios intermedios',
                '16': 'Pago con criptomonedas', '17': 'Otros', '18': 'Devolución',      
                '19': 'Tarjeta de débito','20': 'Dinero electrónico'
            }

            # Procesar el código de la forma de pago
            if forma_pago is not None:
                # Asegurar que el código sea de dos dígitos, con ceros a la izquierda si es necesario
                codigo_forma_pago = forma_pago.zfill(2)
            else:
                # Si no hay forma de pago, usar el valor predeterminado '17' (Otros)
                codigo_forma_pago = '17'

            # Obtener la descripción de la forma de pago desde el diccionario
            descripcion_forma_pago = formas_pago.get(codigo_forma_pago, 'Desconocido')

            pdf.cell(80, 8, f"Forma de Pago: {descripcion_forma_pago}", border=1, ln=1, align='L')
            pdf.cell(80, 8, f"Total con impuestos: ${float(factura['Total con impuestos']):.2f}", border=1, ln=1, align='L')

        pdf.output(archivo_pdf, 'F')
        return archivo_pdf

    # Llamar a la función para generar el PDF
    archivo_pdf = generar_pdf_facturas(facturas_info, productos_info)

    # Asegúrate de tener el DataFrame df_facturas ya definido antes de esta parte
    # Convertir DataFrame a HTML para previsualización
    tabla_html = df_facturas.to_html(classes='preview-table', index=False)


    # Convertir DataFrame a JSON para pasar a JavaScript
    datos_json = df_facturas.to_json(orient='records')

    # Retornar los archivos generados para su descarga
    return render_template(
        'upload.html',
        excel_report=archivo_excel,
        pdf_report=archivo_pdf,
        tabla_html=tabla_html,
        datos_json=datos_json
    )



# Ruta para descargar el archivo Excel
@app.route('/download_excel')
def download_excel():
    filename = request.args.get('filename')
    custom_name = request.args.get('custom_name')
    if not custom_name.endswith('.xlsx'):
        custom_name += '.xlsx'
    return send_file(filename, as_attachment=True, download_name=custom_name)

# Ruta para descargar el archivo Pdf
@app.route('/download_pdf')
def download_pdf():
    filename = request.args.get('filename')
    custom_name = request.args.get('custom_name')
    if not custom_name.endswith('.pdf'):
        custom_name += '.pdf'
    return send_file(filename, as_attachment=True, download_name=custom_name)

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        flash("Debes iniciar sesión para acceder a esta página.")
        return redirect('/login')
    
    # Obtener parámetros de filtro de la URL
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    clientes_filtro = request.args.getlist('clientes[]')
    vendedores_filtro = request.args.getlist('vendedores[]')
    rango_monto = request.args.get('rango_monto')
    
    # Filtrar facturas según los parámetros
    facturas_filtradas = facturas_info.copy()
    
    # Aplicar filtro de fechas
    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
            facturas_filtradas = [
                f for f in facturas_filtradas 
                if fecha_inicio_dt <= datetime.strptime(f['Fecha de Emisión'], '%d/%m/%Y') <= fecha_fin_dt
            ]
        except ValueError:
            flash("Formato de fecha inválido")
    
    # Aplicar filtro de clientes por RUC
    if clientes_filtro:
        facturas_filtradas = [
            f for f in facturas_filtradas
            if f['RUC del Comprador'] in clientes_filtro
        ]

    # Aplicar filtro de vendedores por RUC
    if vendedores_filtro:
        facturas_filtradas = [
            f for f in facturas_filtradas
            if f['RUC del Vendedor'] in vendedores_filtro
        ]

    # Obtener lista única de RUCs de clientes y sus nombres
    clientes_unicos = sorted(list(set(f['RUC del Comprador'] for f in facturas_info)))
    clientes_nombres = {f['RUC del Comprador']: f['Razón Social comprador'] for f in facturas_info}

    # Obtener lista única de RUCs de vendedores y sus nombres
    vendedores_unicos = sorted(list(set(f['RUC del Vendedor'] for f in facturas_info)))
    vendedores_nombres = {f['RUC del Vendedor']: f['Razón Social del Vendedor'] for f in facturas_info}

    
    # Aplicar filtro de rango de monto
    if rango_monto:
        try:
            min_monto, max_monto = map(float, rango_monto.split('-'))
            facturas_filtradas = [
                f for f in facturas_filtradas 
                if min_monto <= float(f['Total con impuestos']) <= max_monto
            ]
        except ValueError:
            flash("Rango de monto inválido")

    # Procesar datos filtrados
    total_facturas = len(facturas_filtradas)
    total_ventas = sum(float(f['Total con impuestos']) for f in facturas_filtradas)
    ventas_sin_impuestos = sum(float(f['Total sin impuestos']) for f in facturas_filtradas)
    promedio_venta = total_ventas / total_facturas if total_facturas > 0 else 0

    # Análisis por vendedor
    ventas_por_vendedor = {}
    for factura in facturas_filtradas:
        vendedor = factura.get('Razón Social del Vendedor', 'Vendedor no especificado')
        monto = float(factura['Total con impuestos'])
        ventas_por_vendedor[vendedor] = ventas_por_vendedor.get(vendedor, 0) + monto

    # Análisis por cliente
    ventas_por_cliente = {}
    for factura in facturas_filtradas:
        cliente = factura.get('Razón Social comprador', 'Cliente no especificado')
        monto = float(factura['Total con impuestos'])
        ventas_por_cliente[cliente] = ventas_por_cliente.get(cliente, 0) + monto

    # Análisis de IVA
    iva_totales = {
        'IVA 0%': sum(f.get('IVA 0%', 0) for f in facturas_filtradas),
        'IVA 5%': sum(f.get('IVA 5%', 0) for f in facturas_filtradas),
        'IVA 12%': sum(f.get('IVA 12%', 0) for f in facturas_filtradas),
        'IVA 15%': sum(f.get('IVA 15%', 0) for f in facturas_filtradas)
    }

    # Análisis temporal
    ventas_por_mes = {}
    for factura in facturas_filtradas:
        fecha = datetime.strptime(factura['Fecha de Emisión'], '%d/%m/%Y')
        mes_año = fecha.strftime('%Y-%m')
        ventas_por_mes[mes_año] = ventas_por_mes.get(mes_año, 0) + float(factura['Total con impuestos'])


    return render_template('dashboard.html',
                         total_facturas=total_facturas,
                         total_ventas=total_ventas,
                         ventas_sin_impuestos=ventas_sin_impuestos,
                         promedio_venta=promedio_venta,
                         ventas_por_cliente=ventas_por_cliente,
                         ventas_por_vendedor=ventas_por_vendedor,
                         iva_totales=iva_totales,
                         ventas_por_mes=ventas_por_mes,
                         clientes_unicos=clientes_unicos,
                         vendedores_unicos=vendedores_unicos,
                         clientes_nombres=clientes_nombres,
                         vendedores_nombres=vendedores_nombres,
                         fecha_inicio=fecha_inicio,
                         fecha_fin=fecha_fin,
                         clientes_filtro=clientes_filtro,
                         vendedores_filtro=vendedores_filtro,
                         rango_monto=rango_monto)

@app.route('/reporte/<int:id>')
def ver_reporte(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    reporte = conn.execute('SELECT * FROM reportes WHERE id = ?', (id,)).fetchone()
    conn.close()

    if reporte is None:
        flash("Reporte no encontrado.")
        return redirect(url_for('reportes_anteriores'))
    
    return render_template('ver_reporte.html', reporte=reporte)

@app.route('/borrar-reportes', methods=['POST'])
def borrar_reportes():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    reportes_ids = request.json.get('ids', [])
    if reportes_ids:
        conn = get_db_connection()
        conn.executemany('DELETE FROM reportes WHERE id = ?', [(id,) for id in reportes_ids])
        conn.commit()
        conn.close()
    
    return '', 204

@app.route('/borrar-reporte/<int:id>', methods=['DELETE'])
def borrar_reporte(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    conn.execute('DELETE FROM reportes WHERE id = ?', (id,))
    conn.commit()
    conn.close()

    return '', 204


if __name__ == '__main__':
    app.run(debug=True)
