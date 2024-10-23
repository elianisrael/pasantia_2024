from flask import Flask, render_template, request, redirect, send_file # type: ignore
import xml.etree.ElementTree as ET
import pandas as pd # type: ignore
import openpyxl # type: ignore
from fpdf import FPDF # type: ignore
import os
import datetime
from openpyxl.styles import Font, PatternFill, Border, Side # type: ignore
import sqlite3
from werkzeug.security import generate_password_hash
from werkzeug.security import check_password_hash

app = Flask(__name__)

# Función para conectar y crear la base de datos
def init_db():
    conn = sqlite3.connect('app.db')  # Nombre del archivo de la base de datos
    cursor = conn.cursor()

    # Crear tabla de usuarios
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            email TEXT NOT NULL,
            password TEXT NOT NULL
        )
    ''')

    # Crear tabla de facturas
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS facturas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            detalles TEXT NOT NULL,
            fecha TEXT NOT NULL,
            total REAL NOT NULL
        )
    ''')

    conn.commit()
    conn.close()

# Llamar a la función para crear las tablas al inicio
init_db()




facturas_info = []

# Ruta para la página principal
@app.route('/')
def index():
    return render_template('inicio.html')
#Ruta para página de inicio
@app.route('/inicio')
def inicio():
    return render_template('inicio.html')

@app.route('/login')
def login():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def handle_login():
    email = request.form['email']
    password = request.form['password']

    conn = sqlite3.connect('app.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM usuarios WHERE email = ?', (email,))
    user = cursor.fetchone()
    conn.close()

    if user and check_password_hash(user[3], password):
        return redirect('/upload')  # Redirigir al dashboard o página principal
    else:
        return 'Error: usuario o contraseña incorrectos'
    
@app.route('/upload')
def upload_1():
    return render_template('upload.html')

@app.route('/facturas')
def ver_facturas():
    conn = sqlite3.connect('app.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM facturas')
    facturas = cursor.fetchall()
    conn.close()

    return render_template('facturas.html', facturas=facturas)

@app.route('/crear_factura', methods=['POST'])
def crear_factura():
    detalles = request.form['detalles']
    fecha = request.form['fecha']
    total = request.form['total']

    conn = sqlite3.connect('app.db')
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO facturas (detalles, fecha, total)
        VALUES (?, ?, ?)
    ''', (detalles, fecha, total))
    conn.commit()
    conn.close()

    return redirect('/facturas')

@app.route('/registro')
def registro():
    return render_template('registro.html')

@app.route('/registro', methods=['POST'])
def handle_registro():
    username = request.form['username']
    email = request.form['email']
    password = request.form['password']

    # Encriptar la contraseña antes de guardarla
    hashed_password = generate_password_hash(password)

    # Guardar en la base de datos
    conn = sqlite3.connect('app.db')
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO usuarios (username, email, password)
        VALUES (?, ?, ?)
    ''', (username, email, hashed_password))
    conn.commit()
    conn.close()

    return redirect('/login')

#Ruta para sección en donde se almacenaran reportes anteriores
@app.route('/reportes.anteriores')
def reportesanteriores():
    return render_template('/reportes.html')
# Ruta para subir los archivos XML y generar los reportes
@app.route('/upload', methods=['GET','POST'])
def upload_files():
    # Obtener los archivos XML subidos por el usuario
    uploaded_files = request.files.getlist('xml_files')
    
    # Crear listas para almacenar la información de las facturas
    # facturas_info = []

    for file in uploaded_files:
        try:
            tree = ET.parse(file)
            root = tree.getroot()

            # Extraer información de la factura (igual que en tu código original)
            estado = root.find('estado').text if root.find('estado') is not None else None
            numero_autorizacion = root.find('numeroAutorizacion').text if root.find('numeroAutorizacion') is not None else None
            fecha_autorizacion = root.find('fechaAutorizacion').text if root.find('fechaAutorizacion') is not None else None
            ambiente = root.find('ambiente').text if root.find('ambiente') is not None else None

            comprobante_element = root.find('comprobante')
            if comprobante_element is not None:
                comprobante_xml = comprobante_element.text.strip()
                if comprobante_xml.startswith('<?xml'):
                    comprobante_xml = comprobante_xml.split('?>', 1)[1].strip()
            else:
                continue

            comprobante_root = ET.fromstring(comprobante_xml)
            info_tributaria = comprobante_root.find('infoTributaria')
            razon_social = info_tributaria.find('razonSocial').text if info_tributaria is not None else None
            ruc = info_tributaria.find('ruc').text if info_tributaria is not None else None
            clave_acceso = info_tributaria.find('claveAcceso').text if info_tributaria is not None else None
            
            codigo_factura = info_tributaria.find('secuencial').text if info_tributaria is not None else None
            info_factura = comprobante_root.find('infoFactura')
            fecha_emision = info_factura.find('fechaEmision').text if info_factura is not None else None
            total_sin_impuestos = info_factura.find('totalSinImpuestos').text if info_factura is not None else None
            importe_total = info_factura.find('importeTotal').text if info_factura is not None else None
            detalles = comprobante_root.find('detalles')

            if detalles is not None:
                productos = detalles.findall('detalle')
                ivas = {"0%": 0, "5%": 0, "12%": 0, "15%": 0}
                total_factura = 0
                for producto in productos:
                    codigo = producto.find('codigoPrincipal').text if producto.find('codigoPrincipal') is not None else None
                    descripcion = producto.find('descripcion').text if producto.find('descripcion') is not None else None
                    cantidad = float(producto.find('cantidad').text) if producto.find('cantidad') is not None else 0.0
                    precio_unitario = float(producto.find('precioUnitario').text) if producto.find('precioUnitario') is not None else 0.0
                    precio_total_sin_impuesto = float(producto.find('precioTotalSinImpuesto').text) if producto.find('precioTotalSinImpuesto') is not None else 0.0

                    impuesto_element = producto.find('impuestos')
                    if impuesto_element is not None:
                        valor_impuesto = impuesto_element.find('impuesto')
                        impuesto = float(valor_impuesto.find('valor').text) if valor_impuesto is not None else 0.0
                        tarifa = float(valor_impuesto.find('tarifa').text) if valor_impuesto.find('tarifa') is not None else 0.0
                    else:
                        impuesto = 0.0
                        tarifa = 0.0
                    total_producto = precio_total_sin_impuesto + impuesto
                    total_factura += total_producto

                     # Determinar el tipo de IVA y sumar al diccionario de IVAs
                    if tarifa == 0:
                         ivas["0%"] += impuesto
                    elif tarifa == 5:
                         ivas["5%"] += impuesto
                    elif tarifa == 12:
                         ivas["12%"] += impuesto
                    elif tarifa == 15:
                        ivas["15%"] += impuesto
                    else:
                          ivas[f"{tarifa}%"] = ivas.get(f"{tarifa}%", 0) + impuesto

            # Crear una cadena con los IVAs de la factura
                    ivas_str = ", ".join([f"{k}: ${v:.2f}" for k, v in ivas.items() if v > 0])

                    facturas_info.append({
                        'Codigo Factura':codigo_factura,
                        'Estado de la autorización': estado,
                        'Fecha de autorización': fecha_autorizacion,
                        'Ambiente': ambiente,
                        'Razón Social': razon_social,
                        'RUC': ruc,
                        'Fecha de Emisión': fecha_emision,
                        'Código': codigo,
                        'Descripción': descripcion,
                        'Cantidad': cantidad,
                        'Precio Unitario': precio_unitario,
                        'iva 0%': ivas.get("0%", 0),
                        'iva 5%': ivas.get("5%", 0),
                        'iva 12%': ivas.get("12%", 0),
                        'iva 15%': ivas.get("15%", 0),
                        'Total sin impuestos': total_sin_impuestos,
                        'Total con impuestos': importe_total,
                        'Número de autorización': numero_autorizacion,
                        'Clave de Acceso': clave_acceso
                    })
        except ET.ParseError:
            continue
        except Exception as e:
            print(f"Error: {e}")

    # Generar reporte en Excel
    df_facturas = pd.DataFrame(facturas_info)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo_excel = f'reporte_facturas_{timestamp}.xlsx'
    #df_facturas.to_excel(archivo_excel, index=False)

    writer = pd.ExcelWriter(archivo_excel, engine='openpyxl')

    # Guardar el DataFrame en Excel
    df_facturas.to_excel(writer, index=False, sheet_name='Reporte')

    # Obtener la hoja de trabajo
    workbook = writer.book
    worksheet = workbook['Reporte']

    # Aplicar formato a la hoja
    for idx, col in enumerate(df_facturas.columns):
        # Establecer el ancho de la columna basado en la longitud máxima en la columna
        max_len = max(df_facturas[col].astype(str).map(len).max(), len(col))
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx+1)].width = max_len + 2

    # Aplicar estilo a la fila de encabezados
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')

    # Aplicar bordes a todas las celdas
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border

    # Aplicar filtro
    worksheet.auto_filter.ref = worksheet.dimensions

    # Congelar la primera fila
    worksheet.freeze_panes = 'A2'

    # Guardar el archivo
    writer.close()

    archivo_pdf = f'reporte_facturas_{timestamp}.pdf'
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_font('Arial', '', 10)

    for i, factura in enumerate(facturas_info):
        pdf.add_page()
    
        # Título
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, 'Factura', 0, 1, 'C')
        pdf.ln(5)
    
        # Información del emisor
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, f"Razón Social: {factura['Razón Social']}", 0, 1)
        pdf.cell(0, 10, f"RUC: {factura['RUC']}", 0, 1)
        pdf.ln(5)
    
        # Información de la factura
        pdf.set_font('Arial', '', 10)
        pdf.cell(0, 10, f"Fecha de Emisión: {factura['Fecha de Emisión']}", 0, 1)
        pdf.cell(0, 10, f"Número de Autorización: {factura['Número de autorización']}", 0, 1)
        pdf.cell(0, 10, f"Clave de Acceso: {factura['Clave de Acceso']}", 0, 1)
        pdf.ln(5)
    
        # Detalles de la factura
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, "Detalles de la Factura", 0, 1)
        pdf.ln(2)
    
        pdf.set_font('Arial', '', 10)
        pdf.cell(30, 10, "Código", 1)
        pdf.cell(60, 10, "Descripción", 1)
        pdf.cell(30, 10, "Cantidad", 1)
        pdf.cell(30, 10, "Precio Unitario", 1)
        pdf.cell(40, 10, "Total", 1)
        pdf.ln()
    
        pdf.cell(30, 10, str(factura['Código']), 1)
        pdf.cell(60, 10, factura['Descripción'], 1)
        pdf.cell(30, 10, str(factura['Cantidad']), 1)
        pdf.cell(30, 10, f"${factura['Precio Unitario']:.2f}", 1)
        total = factura['Cantidad'] * factura['Precio Unitario']
        pdf.cell(40, 10, f"${total:.2f}", 1)
        pdf.ln(15)
    
        # Totales
        pdf.cell(120)
        pdf.cell(30, 10, "Subtotal:", 0)
        pdf.cell(40, 10, f"${float(factura['Total sin impuestos']):.2f}", 0)
        pdf.ln()
        pdf.cell(120)
        pdf.cell(30, 10, "IVA 0%:", 0)
        pdf.cell(40, 10, f"${factura['iva 0%']:.2f}", 0)
        pdf.ln()
        pdf.cell(120)
        pdf.cell(30, 10, "IVA 5%:", 0)
        pdf.cell(40, 10, f"${factura['iva 5%']:.2f}", 0)
        pdf.ln()
        pdf.cell(120)
        pdf.cell(30, 10, "IVA 12%:", 0)
        pdf.cell(40, 10, f"${factura['iva 12%']:.2f}", 0)
        pdf.ln()
        pdf.cell(120)
        pdf.cell(30, 10, "IVA 15%:", 0)
        pdf.cell(40, 10, f"${factura['iva 15%']:.2f}", 0)
        pdf.ln()
        pdf.cell(120 )
        pdf.cell(30, 10, "Total:", 0)
        pdf.cell(40, 10, f"${float(factura['Total con impuestos']):.2f}", 0)
        pdf.ln(10)

    pdf.output(archivo_pdf, 'F')

    # Convertir DataFrame a HTML para previsualización
    tabla_html = df_facturas.to_html(classes='preview-table', index=False)

    # Convertir DataFrame a JSON para pasar a JavaScript
    datos_json = df_facturas.to_json(orient='records')

    # Retornar los archivos generados para su descarga
    return render_template('upload.html', excel_report=archivo_excel, pdf_report=archivo_pdf,tabla_html=tabla_html,
                           datos_json=datos_json)



# Ruta para descargar el archivo Excel
@app.route('/download_excel')
def download_excel():
    filename = request.args.get('filename')
    custom_name = request.args.get('custom_name')
    if not custom_name.endswith('.xlsx'):
        custom_name += '.xlsx'
    return send_file(filename, as_attachment=True, download_name=custom_name)

# Ruta para descargar el archivo Pdf
@app.route('/download_pdf')
def download_pdf():
    filename = request.args.get('filename')
    custom_name = request.args.get('custom_name')
    if not custom_name.endswith('.pdf'):
        custom_name += '.pdf'
    return send_file(filename, as_attachment=True, download_name=custom_name)

@app.route('/dashboard')
def dashboard():
    # Procesa los datos de las facturas (asumiendo que tienes acceso a facturas_info)
    total_facturas = len(facturas_info)
    total_ventas = sum(float(factura['Total con impuestos']) for factura in facturas_info)
    ventas_sin_impuestos = sum(float(factura['Total sin impuestos']) for factura in facturas_info)
    promedio_venta = total_ventas / total_facturas if total_facturas > 0 else 0
    
    # Datos para gráficos
    ventas_por_cliente = {}
    
    for factura in facturas_info:
        cliente = factura['Razón Social']
        monto = float(factura['Total con impuestos'])
        ventas_por_cliente[cliente] = ventas_por_cliente.get(cliente, 0) + monto

    iva_totales = {
        'IVA 0%': sum(factura['iva 0%'] for factura in facturas_info),
        'IVA 5%': sum(factura['iva 5%'] for factura in facturas_info),
        'IVA 12%': sum(factura['iva 12%'] for factura in facturas_info),
        'IVA 15%': sum(factura['iva 15%'] for factura in facturas_info)
    }

    return render_template('dashboard.html', 
                           total_facturas=total_facturas,
                           total_ventas=total_ventas,
                           ventas_sin_impuestos=ventas_sin_impuestos,
                           promedio_venta=promedio_venta,
                           ventas_por_cliente=ventas_por_cliente,
                           iva_totales=iva_totales)  

if __name__ == '__main__':
    app.run(debug=True)
